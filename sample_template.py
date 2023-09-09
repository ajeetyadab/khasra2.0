import openpyxl
file1="C:/Users/acer/Desktop/sample.xlsx"

fasal_values=["धान - अधिक उपज वाला","दाल उर्द","मक्का","अन्य तरकारियां","ईख बोई हुई" ,"ईख पेड़ी","अन्य चारा","रिक्त"]
sichai_vidhi=["बोरिंग निजी (डीजल)","बोरिंग निजी (विधुत)","अन्य"]
sinchit_asinchit=["sichitArea","asichitArea"]
gram_samaj=["आबादी","खाद के गड्डे","चकमार्ग","नदी","नवीन परती","रास्ता","रेत"]

value=0

wb=openpyxl.load_workbook(file1)
ws=wb["Sheet1"]

rows=ws.max_row
column=ws.max_column
print(rows)
print(column)

def min_rakba_condition(r):
    global value
    if ws.cell(r,1).value==ws.cell(r+1,1).value:
        value+=1
        print(value)
        ws.cell(r,6).value=f"ksn-{value-1}"
    
    elif r>1 and ws.cell(r,1).value==ws.cell(r-1,1).value:
        value+=1
        ws.cell(r,6).value=f"ksn-{value-1}"
        
        
    else:
        value=0
        ws.cell(r,6).value=f"ksn-{value}"
        

for r in range(1,rows+1):
    #ws.cell(r,6).value="ksn-0"
    if ws.cell(r,3).value>0.7:
        ws.cell(r,8).value=sichai_vidhi[0]
    else:
        ws.cell(r,8).value=sichai_vidhi[2]
     
    ws.cell(r,9).value=sinchit_asinchit[0]
     
    if ws.cell(r,5).value in gram_samaj:
        ws.cell(r,7).value=fasal_values[7]
    else:
        ws.cell(r,7).value=fasal_values[0]
        
    min_rakba_condition(r)           
    



wb.save(file1)     
        
    
