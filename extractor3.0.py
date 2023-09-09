from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import TimeoutException,NoAlertPresentException,UnexpectedAlertPresentException
from selenium.webdriver.firefox.service import Service
from selenium.webdriver.firefox.options import Options
import time
import openpyxl
import os


print(os.getcwd())
file1="C:/Users/acer/Desktop/sample.xlsx"
file2="C:/Users/acer/Desktop/data.xlsx"
workbook1=openpyxl.load_workbook(file1)
workbook2=openpyxl.load_workbook(file2)
sheet1=workbook1["Sheet1"]
sheet2=workbook2["credentials"]


halka_name=sheet2.cell(2,1).value
pass_word=sheet2.cell(2,2).value
gram_name=sheet2.cell(2,3).value
total_gata=sheet2.cell(2,4).value
start_gata=sheet2.cell(2,5).value




options = Options()
driver_path = "./chromedriver.exe"
driver = webdriver.Chrome()




#ops=Options()
#ops.headless=False
#ser_ob = Service(r"C:/Users/acer/Desktop/python codes/geckodriver.exe")
#driver = webdriver.Firefox(service=ser_ob,options=ops)






#TEHSHIL_VALUE = "00723"





number_x_path_map = {
    "1": "//*[@id=\"searchGata\"]/div/div[3]/table/tbody/tr[1]/td[1]/a/div",
    "2": "//*[@id=\"searchGata\"]/div/div[3]/table/tbody/tr[1]/td[2]/a/div",
    "3": "//*[@id=\"searchGata\"]/div/div[3]/table/tbody/tr[1]/td[3]/a/div",
    "4": "//*[@id=\"searchGata\"]/div/div[3]/table/tbody/tr[1]/td[4]/a/div",
    "5": "//*[@id=\"searchGata\"]/div/div[3]/table/tbody/tr[2]/td[1]/a/div",
    "6": "//*[@id=\"searchGata\"]/div/div[3]/table/tbody/tr[2]/td[2]/a/div",
    "7": "//*[@id=\"searchGata\"]/div/div[3]/table/tbody/tr[2]/td[3]/a/div",
    "8": "//*[@id=\"searchGata\"]/div/div[3]/table/tbody/tr[2]/td[4]/a/div",
    "9": "//*[@id=\"searchGata\"]/div/div[3]/table/tbody/tr[3]/td[1]/a/div",
    "0": "//*[@id=\"searchGata\"]/div/div[3]/table/tbody/tr[3]/td[2]/a/div",
    "delete": "//*[@id=\"searchGata\"]/div/div[3]/table/tbody/tr[3]/td[3]/a/div",
    "clear": "//*[@id=\"searchGata\"]/div/div[3]/table/tbody/tr[3]/td[4]/a/div"
}

def load_first_page():
    driver.get("http://164.100.59.148/")
    time.sleep(2)
    #driver.find_element(By.XPATH,"/html/body/center/main/div/div/ul/li[4]/a/div/div[1]").click()
    driver.find_element(By.XPATH,"/html/body/center/main/div/div/ul/li[5]/a/div").click()
    time.sleep(1)






    
def load_second_page():
    selectDistrict = Select(driver.find_element(By.ID, "up_district"))
    selectDistrict.select_by_visible_text("रामपुर")
    time.sleep(1)
    selectTehsil = Select(driver.find_element(By.ID, "up_tehsil"))
    selectTehsil.select_by_visible_text("स्वार")
    
    
    time.sleep(.5)
    selecthalka = Select(driver.find_element(By.ID, "up_halka"))
    time.sleep(.5)
    selecthalka.select_by_visible_text(halka_name)
    captcha_value = driver.find_element(By.ID, "CaptchaDiv").text
    driver.find_element(By.ID, "CaptchaInput").send_keys(captcha_value)
    driver.find_element(By.ID, "password").send_keys(pass_word)
    time.sleep(1)
    driver.find_element(By.CLASS_NAME, "login100-form-btn").click()



def load_third_page():
    time.sleep(3)
    Select(driver.find_element(By.ID,"gram_name")).select_by_visible_text(gram_name)
    time.sleep(1)
    Select(driver.find_element(By.ID,"fasalYear")).select_by_visible_text("1430 (1 जुलाई 2022 से 30 जून 2023)")
    #Select(driver.find_element(By.ID,"fasalYear")).select_by_visible_text("1428 (1 जुलाई 2020 से 30 जून 2021)")
    time.sleep(3)
    Select(driver.find_element(By.ID,"fasal")).select_by_visible_text("खरीफ की फसल (10 अगस्त से 30 सितम्बर)")
    time.sleep(1)
    alert_window_0 = driver.switch_to.alert
    print(alert_window_0.text)
    alert_window_0.accept()
    driver.find_element(By.CLASS_NAME, "login100-form-btn").click()
    
def load_fourth_page():
    time.sleep(3)
    alert_window = driver.switch_to.alert
    print(alert_window.text)
    alert_window.accept()
    time.sleep(2)
    
    #driver.find_element(By.XPATH, "//*[@id=\"link2\"]/a").click()#this element is not working in fiefox
    driver.find_element(By.XPATH,"//*[@id=\"link2\"]").click()
    fill_form()
    

def fill_form():
    time.sleep(0.5)
    for i in range(start_gata,total_gata):
        fill_khasra_pravisti(i)
        
        

def click_digits(digits):
    for digit in digits:
        driver.find_element(By.XPATH, number_x_path_map[digit]).click()
        
        
def search_number(number):
    click_digits(str(number))
    driver.find_element(By.XPATH, "//*[@id=\"sgw\"]/button/i").click()
    
    
    

def fill_khasra_pravisti(i):
    search_number(i)
    time.sleep(1.5)
    gata_element_list=driver.find_elements(By.XPATH,"//input[@name=\"khata_number\"]")
    if len(gata_element_list)==0:
        print("gata no found error", f"{i}")
        #driver.find_element(By.XPATH,"//*[@id=\"searchGata\"]/div/div[3]/table/tbody/tr[3]/td[4]/a").click()#link not working in firefox
        driver.find_element(By.XPATH,"/html/body/center/main/div/div[1]/div/div[3]/table/tbody/tr[3]/td[4]/a/div").click()
        time.sleep(1)
     
    elif len(gata_element_list)==1:
        gata_element_list[0].click()
        khatauni_detail_extracion(i)
         
         
    elif len(gata_element_list)>1:
        for gata_index in range(0,len(gata_element_list)):
                driver.find_element(By.XPATH,"/html/body/center/main/div/div[1]/div/div[3]/table/tbody/tr[3]/td[4]").click()
                search_number(i)
                time.sleep(2)
                gata_element_list=driver.find_elements(By.XPATH,"//input[@name=\"khata_number\"]") # if gata_element_list not updated here it will through stale data error 
                gata_element_list[gata_index].click()
                #driver.find_element(By.ID,f"ksn-{gata_index}").click()
                khatauni_detail_extracion(i)
                if gata_index==len(gata_element_list)-1:
                    break
                #else:
                 #   search_number(i)
                 #   if len(driver.find_elements(By.XPATH,"//input[@name=\"khata_number\"]"))==0:# this line is being added to correct list index out of range error 
                  #      search_number(i)
                  #      time.sleep(1)
                
            
    


def khatauni_detail_extracion(i):
    driver.find_element(By.XPATH,"//*[@id=\"case_frm\"]/button[2]").click()
    time.sleep(1)
    gata_num_list=(driver.find_element(By.XPATH,"//*[@id=\"tabs-container\"]/div[1]/label[1]").text).split(" : ")
    print(gata_num_list)
    khata_name_list=(driver.find_element(By.XPATH,"//*[@id=\"tabs-container\"]/div[2]").text).split(" : ")
    gata_uid_list=(driver.find_element(By.XPATH,"//*[@id=\"tabs-container\"]/div[1]/label[2]").text).split(" : ")
    gata_area_list=(driver.find_element(By.XPATH,"//*[@id=\"tabs-container\"]/div[1]/label[3]").text).split(" : ")
    gata_area=gata_area_list[1].strip()
    
    sheet1.append([f'{i}',f'{gata_num_list[1].strip()}',float(f'{gata_area}'),f'{gata_uid_list[1].strip()}',f'{khata_name_list[1].strip()}'])
    #print(sheet1.cell(i,1).value,sheet1.cell(i,2).value,type(sheet1.cell(i,3).value))
    workbook1.save(file1)
    
    driver.find_element(By.XPATH,"//*[@id=\"content\"]/center/header/div/div[7]/div").click()                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                           
    time.sleep(1)
    
    

    

    


load_first_page()
load_second_page()
load_third_page()
load_fourth_page()
workbook1.save(file1)


    

