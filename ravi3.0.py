from selenium import webdriver
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import TimeoutException,NoAlertPresentException,UnexpectedAlertPresentException,NoSuchElementException
from selenium.webdriver.firefox.service import Service
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.action_chains import ActionChains
import time
import openpyxl




file1="C:/Users/acer/Desktop/hrd.xlsx"
file2="C:/Users/acer/Desktop/data.xlsx"
workbook1=openpyxl.load_workbook(file1)
workbook2=openpyxl.load_workbook(file2)
sheet1=workbook1["analysis_copy"]
sheet2=workbook2["credentials"]




halka_name=sheet2.cell(2,1).value
pass_word=sheet2.cell(2,2).value
gram_name=sheet2.cell(2,3).value
total_gata=sheet2.cell(2,4).value
start_gata=sheet2.cell(2,5).value


ttime = time.localtime()





serv_obj=Service("chromedriver.exe")
driver=webdriver.Chrome(service=serv_obj)
actions=ActionChains(driver)
mywait=WebDriverWait(driver,10)



row_count=sheet1.max_row
col_count=sheet1.max_column




iterable_gata=[]
iterable_id=[]
iterable_fasal=[]
iterable_boring=[]
iterable_sinchit_asinchit=[]
iterable_area=[]

for i in range(1,row_count+1):
    iterable_gata.append(sheet1.cell(i+1,2).value)
    iterable_id.append(sheet1.cell(i+1,7).value)
    iterable_fasal.append(sheet1.cell(i+1,11).value)
    iterable_boring.append(sheet1.cell(i+1,9).value)
    iterable_sinchit_asinchit.append(sheet1.cell(i+1,10).value)
    iterable_area.append(sheet1.cell(i+1,4).value)

    
print(iterable_gata)
print(iterable_id)
print(iterable_fasal)
print(iterable_boring)
print(iterable_sinchit_asinchit)
print(iterable_area)
    

#ops=Options()
#ops.headless=False
#ser_ob = Service(r"C:/Users/acer/Desktop/python codes/geckodriver.exe")
#driver = webdriver.Firefox(service=ser_ob,options=ops)






#TEHSHIL_VALUE = "00723"





number_x_path_map = {
    "1": "//*[@id=\"searchGata\"]/div/div[3]/table/tbody/tr[1]/td[1]/a/div",
    "2": "//*[@id=\"searchGata\"]/div/div[3]/table/tbody/tr[1]/td[2]/a/div",
    "3": "//*[@id=\"searchGata\"]/div/div[3]/table/tbody/tr[1]/td[3]/a/div",
    "4": "//*[@id=\"searchGata\"]/div/div[3]/table/tbody/tr[1]/td[4]/a/div",
    "5": "//*[@id=\"searchGata\"]/div/div[3]/table/tbody/tr[2]/td[1]/a/div",
    "6": "//*[@id=\"searchGata\"]/div/div[3]/table/tbody/tr[2]/td[2]/a/div",
    "7": "//*[@id=\"searchGata\"]/div/div[3]/table/tbody/tr[2]/td[3]/a/div",
    "8": "//*[@id=\"searchGata\"]/div/div[3]/table/tbody/tr[2]/td[4]/a/div",
    "9": "//*[@id=\"searchGata\"]/div/div[3]/table/tbody/tr[3]/td[1]/a/div",
    "0": "//*[@id=\"searchGata\"]/div/div[3]/table/tbody/tr[3]/td[2]/a/div",
    "delete": "//*[@id=\"searchGata\"]/div/div[3]/table/tbody/tr[3]/td[3]/a/div",
    "clear": "//*[@id=\"searchGata\"]/div/div[3]/table/tbody/tr[3]/td[4]/a/div"
}

def load_first_page():
    driver.get("http://164.100.59.148/")
    time.sleep(1)
    driver.find_element(By.XPATH,"//*[@id=\"about_us\"]/div/div[2]/a").click()
    time.sleep(2)
    if ttime.tm_hour >=9 and ttime.tm_hour <= 16:
        driver.find_element(By.XPATH,"/html/body/center/main/div/div/ul/li[3]/a/div/div[1]").click() # 8- 5 pm link
        
		
    else:
        driver.find_element(By.XPATH,"/html/body/center/main/div/div/ul/li[5]/a/div/div[1]").click() # after 8 pm link
        
    
    time.sleep(1)






    
def load_second_page():
    selectDistrict = Select(driver.find_element(By.ID, "up_district"))
    selectDistrict.select_by_visible_text("रामपुर")
    time.sleep(1)
    selectTehsil = Select(driver.find_element(By.ID, "up_tehsil"))
    selectTehsil.select_by_visible_text("स्वार")
    
    
    time.sleep(.5)
    selecthalka = Select(driver.find_element(By.ID, "up_halka"))
    time.sleep(.5)
    selecthalka.select_by_visible_text(halka_name)
    #captcha_value = driver.find_element(By.ID, "CaptchaDiv").text
    #driver.find_element(By.ID, "CaptchaInput").send_keys(captcha_value)
    driver.find_element(By.ID, "password").send_keys(pass_word)
    time.sleep(15)
    driver.find_element(By.CLASS_NAME, "login100-form-btn").click()



def load_third_page():
    time.sleep(3)
    Select(driver.find_element(By.ID,"gram_name")).select_by_visible_text(gram_name)
    time.sleep(1)
    
    Select(driver.find_element(By.ID,"fasalYear")).select_by_visible_text("1430 (1 जुलाई 2022 से 30 जून 2023)")
    
    
    time.sleep(3)
    #Select(driver.find_element(By.ID,"fasal")).select_by_visible_text("खरीफ की फसल (10 अगस्त से 30 सितम्बर)") # FOR KHAREEF
    Select(driver.find_element(By.ID,"fasal")).select_by_visible_text("रबी की फसल (1 जनवरी से 28 फरवरी)") # FOR RABI
    
    time.sleep(1)
    alert_window_0 = driver.switch_to.alert
    print(alert_window_0.text)
    alert_window_0.accept()
    driver.find_element(By.CLASS_NAME, "login100-form-btn").click()
    
def load_fourth_page():
    time.sleep(3)
    alert_window = driver.switch_to.alert
    print(alert_window.text)
    alert_window.accept()
    time.sleep(2)
    
    #driver.find_element(By.XPATH, "//*[@id=\"link2\"]/a").click()#this element is not working in fiefox
    driver.find_element(By.XPATH,"//*[@id=\"link2\"]").click()
    fill_form()
    

def fill_form():
    time.sleep(0.5)
    for i in range(start_gata,total_gata):
        fill_khasra_pravisti(i)
        
        

def click_digits(digits):
    for digit in digits:
        driver.find_element(By.XPATH, number_x_path_map[digit]).click()
        
        
def search_number(number):
    click_digits(number)
    driver.find_element(By.XPATH, "//*[@id=\"sgw\"]/button/i").click()
    time.sleep(1)
    
    
    

def fill_khasra_pravisti(i):
    print(i)
    search_number(str(iterable_gata[i-1]))
    print(iterable_gata[i-1])
    
    try :
        mywait.until(expected_conditions.presence_of_element_located((By.ID,f"{iterable_id[i-1]}")))
        driver.find_element(By.ID,f"{iterable_id[i-1]}").click()
        print(iterable_id[i-1])
        time.sleep(1)
        driver.find_element(By.XPATH,"//*[@id=\"case_frm\"]/button[2]").click()
        time.sleep(.5)
        Select(driver.find_element(By.ID, "fasal_name")).select_by_visible_text(f"{iterable_fasal[i-1]}")
    except :
         time.sleep(2)
         print("error occued on entry" f"{i}")
         driver.refresh()
         
         driver.find_element(By.XPATH,"//*[@id=\"searchGata\"]/div/div[3]/table/tbody/tr[3]/td[4]/a").click()
         time.sleep(1)
         
         
         search_number(iterable_gata[i-1])
         print(iterable_id[i-1])
         #mywait.until(expected_conditions.presence_of_element_located(By.ID,f"{iterable_id[i-1]}"))
         driver.find_element(By.ID,f"{iterable_id[i-1]}").click()
         print(iterable_id[i-1])
         time.sleep(.5)
         driver.find_element(By.XPATH,"//*[@id=\"case_frm\"]/button[2]").click()
         time.sleep(.5)
         Select(driver.find_element(By.ID, "fasal_name")).select_by_visible_text(f"{iterable_fasal[i-1]}")
         
    
    
    if iterable_fasal[i-1] !="रिक्त":
        if iterable_sinchit_asinchit[i-1] =="sichitArea":
            Select(driver.find_element(By.ID, "agriTech")).select_by_visible_text(f"{iterable_boring[i-1]}")
            driver.find_element(By.ID, "sichitArea").clear()
            driver.find_element(By.ID, "sichitArea").send_keys(iterable_area[i-1])
            print(iterable_area[i-1])
            driver.find_element(By.XPATH,"//*[@id=\"submitBtn\"]").click()
            time.sleep(1)
            driver.find_element(By.XPATH,"//*[@id=\"content\"]/center/header/div/div[7]/div").click()
        else:
            Select(driver.find_element(By.ID, "agriTech")).select_by_visible_text(f"{iterable_boring[i-1]}")
            driver.find_element(By.ID, "asichitArea").clear()
            driver.find_element(By.ID, "asichitArea").send_keys(iterable_area[i-1])
            print(iterable_area[i-1])
            driver.find_element(By.XPATH,"//*[@id=\"submitBtn\"]").click()
            time.sleep(1)
            driver.find_element(By.XPATH,"//*[@id=\"content\"]/center/header/div/div[7]/div").click()
             
    else:
        driver.find_element(By.XPATH,"//*[@id=\"submitBtn\"]").click()
        time.sleep(1)
        driver.find_element(By.XPATH,"//*[@id=\"content\"]/center/header/div/div[7]/div").click()
        
        
        
    
    
    
    
    
        
        
    
    

    
            
    


def khatauni_detail_extracion(i):
    driver.find_element(By.XPATH,"//*[@id=\"case_frm\"]/button[2]").click()
    time.sleep(1)
    
    
    driver.find_element(By.XPATH,"//*[@id=\"content\"]/center/header/div/div[7]/div").click()                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                           
    time.sleep(1)
    
    

    

    


load_first_page()
load_second_page()
load_third_page()
load_fourth_page()



    



