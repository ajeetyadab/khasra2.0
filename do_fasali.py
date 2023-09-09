from selenium import webdriver
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import TimeoutException
import time
import openpyxl
from xutility import read_data

file="C:/Users/acer/Desktop/pgm2.xlsx"
workbook=openpyxl.load_workbook(file, keep_vba=True, data_only=True)

fasal="khareef"
sheet1=workbook[fasal]

row=sheet1.max_row
col=sheet1.max_column
gata_list=[]
print(row,col)
for r in range(1,row+1):
    
    gata_no=str(sheet1.cell(r+1,2).value)
    gata_list.append(gata_no)
    
    
id_list=[]
do_fasli_s=[]
do_fasli_a=[]
uncul_area=[]
gata_area=[]

for r in range(1,row):
    id_value=sheet1.cell(r+1,11).value
    """enter above the id element value from excel sheet"""
    
    id_list.append(id_value)
    do_cell_value_sinchit=sheet1.cell(r+1,14).value
    do_cell_value_asinchit=sheet1.cell(r+1,15).value
    uncul_cell_value=sheet1.cell(r+1,16).value
    gata_area_cell_value=sheet1.cell(r+1,4).value
    
    
    do_fasli_s.append(do_cell_value_sinchit)
    do_fasli_a.append(do_cell_value_asinchit)
    uncul_area.append(uncul_cell_value)
    gata_area.append(gata_area_cell_value)
   
#print(len(do_fasli_s),do_fasli_s)
#print(len(do_fasli_a),do_fasli_a)
#print(len(uncul_area))
print(len(gata_area))

  
#print(gata_list)
#print(id_list)
#print(uncul_area)
print(gata_area)



driver_path = "./chromedriver.exe"
driver = webdriver.Chrome(driver_path)

PASSWORD = "@jIt4hero"
DISTRICT_VALUE = "136"
TEHSHIL_VALUE = "00723"
HALKA_VALUE = "0113600723036"




number_x_path_map = {
    "1": "//*[@id=\"searchGata\"]/div/div[3]/table/tbody/tr[1]/td[1]/a/div",
    "2": "//*[@id=\"searchGata\"]/div/div[3]/table/tbody/tr[1]/td[2]/a/div",
    "3": "//*[@id=\"searchGata\"]/div/div[3]/table/tbody/tr[1]/td[3]/a/div",
    "4": "//*[@id=\"searchGata\"]/div/div[3]/table/tbody/tr[1]/td[4]/a/div",
    "5": "//*[@id=\"searchGata\"]/div/div[3]/table/tbody/tr[2]/td[1]/a/div",
    "6": "//*[@id=\"searchGata\"]/div/div[3]/table/tbody/tr[2]/td[2]/a/div",
    "7": "//*[@id=\"searchGata\"]/div/div[3]/table/tbody/tr[2]/td[3]/a/div",
    "8": "//*[@id=\"searchGata\"]/div/div[3]/table/tbody/tr[2]/td[4]/a/div",
    "9": "//*[@id=\"searchGata\"]/div/div[3]/table/tbody/tr[3]/td[1]/a/div",
    "0": "//*[@id=\"searchGata\"]/div/div[3]/table/tbody/tr[3]/td[2]/a/div",
    "delete": "//*[@id=\"searchGata\"]/div/div[3]/table/tbody/tr[3]/td[3]/a/div",
    "clear": "//*[@id=\"searchGata\"]/div/div[3]/table/tbody/tr[3]/td[4]/a/div"
}


def load_first_page():
    driver.get("http://164.100.59.148/")
    time.sleep(2)
    driver.find_element(By.XPATH,"/html/body/center/main/div/div/ul/li[4]/a/div/div[1]").click()
    time.sleep(.25)
    
    

    
def load_second_page():
    selectDistrict = Select(driver.find_element(By.ID, "up_district"))
    selectDistrict.select_by_visible_text("रामपुर")
    time.sleep(.5)
    selectTehsil = Select(driver.find_element(By.ID, "up_tehsil"))
    selectTehsil.select_by_value(TEHSHIL_VALUE)
    time.sleep(.5)
    selectTehsil = Select(driver.find_element(By.ID, "up_halka"))
    selectTehsil.select_by_value(HALKA_VALUE)
    time.sleep(.5)
    captcha_value = driver.find_element(By.ID, "CaptchaDiv").text
    driver.find_element(By.ID, "CaptchaInput").send_keys(captcha_value)
    driver.find_element(By.ID, "password").send_keys(PASSWORD)
    driver.find_element(By.CLASS_NAME, "login100-form-btn").click()



def load_third_page():
    time.sleep(20)
    # driver.find_element(By.CLASS_NAME, "login100-form-btn").click()
    
    
def load_fourth_page():
    time.sleep(12)
    # driver.find_element(By.XPATH, "//*[@id=\"link2\"]/a/div/div[2]").click()
    fill_form()


def fill_form():
    time.sleep(0.5)
    for i in range(279,285):
    #for i in range(1,len(gata_list)+1):
        fill_khasra_pravisti(gata_list[i-1],i)
        print(gata_list[i-1])
        

def click_digits(digits):
    for digit in digits:
        driver.find_element(By.XPATH, number_x_path_map[digit]).click()
        
        
def search_number(number):
    click_digits((number))
    driver.find_element(By.XPATH, "//*[@id=\"sgw\"]/button/i").click()
    
    
    

def fill_khasra_pravisti(gatalist,iteration):
    search_number(gatalist)
    #time.sleep()
    WebDriverWait(driver, 5).until(expected_conditions.presence_of_element_located((By.NAME, "khata_number")))
    driver.find_element(By.ID,id_list[iteration-1]).click()
    time.sleep(.5)
    driver.find_element(By.XPATH,"//*[@id=\"case_frm\"]/button[7]").click() 
    if do_fasli_s[iteration-1]!=None:
        driver.find_element(By.XPATH,"//*[@id=\"doFasliSichitArea\"]").clear()
        driver.find_element(By.XPATH,"//*[@id=\"doFasliSichitArea\"]").send_keys(do_fasli_s[iteration-1])
        
    elif do_fasli_a[iteration-1]!=None :
        driver.find_element(By.XPATH,"//*[@id=\"doFasliAsichitArea\"]").clear()
        driver.find_element(By.XPATH,"//*[@id=\"doFasliAsichitArea\"]").send_keys(do_fasli_a[iteration-1])
         
    elif uncul_area[iteration-1]!=None:
        Select(driver.find_element(By.XPATH,"//*[@id=\"akrishit_type\"]")).select_by_visible_text(uncul_area[iteration-1])
        driver.find_element(By.XPATH,"//*[@id=\"akrishit_area\"]").clear()
        driver.find_element(By.XPATH,"//*[@id=\"akrishit_area\"]").send_keys(gata_area[iteration-1])
    
    driver.find_element(By.XPATH,"//*[@id=\"tab-4\"]/form/p/table[2]/tbody/tr/td[1]/input[1]").click()
    time.sleep(3)
    driver.find_element(By.XPATH,"//*[@id=\"content\"]/center/header/div/div[7]/div").click()
        
        
        
   


load_first_page()
load_second_page()
load_third_page()
load_fourth_page()



    


