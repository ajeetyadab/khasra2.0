from selenium import webdriver
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import TimeoutException
import time
import openpyxl

file="C:/Users/acer/Desktop/khasra2.0/khtauni.xlsx"
workbook=openpyxl.load_workbook(file)
sheet1=workbook["khareef"]


row=sheet1.max_row
col=sheet1.max_column
gata_list=[]
print(row,col)
for r in range(1,row+1):
    
    gata_no=str(sheet1.cell(r+1,2).value)
    gata_list.append(gata_no)
    
    
id_list=[]


for r in range(1,row):
    id_value=sheet1.cell(r+1,8).value
    id_list.append(id_value)
    
driver_path = "./chromedriver.exe"
driver = webdriver.Chrome(driver_path)

PASSWORD = "@jIt4hero"
DISTRICT_VALUE = "136"
TEHSHIL_VALUE = "00723"
HALKA_VALUE = "0113600723036"




number_x_path_map = {
    "1": "/html/body/center/main/div/div[1]/div[3]/center/div[2]/div[3]/table/tbody/tr[1]/td[1]/a",
    "2": "/html/body/center/main/div/div[1]/div[3]/center/div[2]/div[3]/table/tbody/tr[1]/td[2]/a",
    "3": "/html/body/center/main/div/div[1]/div[3]/center/div[2]/div[3]/table/tbody/tr[1]/td[3]/a",
    "4": "/html/body/center/main/div/div[1]/div[3]/center/div[2]/div[3]/table/tbody/tr[1]/td[4]/a",
    "5": "/html/body/center/main/div/div[1]/div[3]/center/div[2]/div[3]/table/tbody/tr[2]/td[1]/a",
    "6": "/html/body/center/main/div/div[1]/div[3]/center/div[2]/div[3]/table/tbody/tr[2]/td[2]/a",
    "7": "/html/body/center/main/div/div[1]/div[3]/center/div[2]/div[3]/table/tbody/tr[2]/td[3]/a",
    "8": "/html/body/center/main/div/div[1]/div[3]/center/div[2]/div[3]/table/tbody/tr[2]/td[4]/a",
    "9": "/html/body/center/main/div/div[1]/div[3]/center/div[2]/div[3]/table/tbody/tr[3]/td[1]/a",
    "0": "/html/body/center/main/div/div[1]/div[3]/center/div[2]/div[3]/table/tbody/tr[3]/td[2]/a",
    "delete": "//*[@id=\"searchGata\"]/div/div[3]/table/tbody/tr[3]/td[3]/a/div",
    "clear": "//*[@id=\"searchGata\"]/div/div[3]/table/tbody/tr[3]/td[4]/a/div"
}


def load_first_page():
    driver.get("https://upbhulekh.gov.in/public/public_ror/Public_ROR.jsp")
    time.sleep(2)
    driver.find_element(By.XPATH,"//*[@id=\"district\"]/li[57]/a/div").click()
    time.sleep(1)
    driver.find_element(By.XPATH,"//*[@id=\"tehsil\"]/li[6]/a/div").click()
    time.sleep(1)
    driver.find_element(By.XPATH,"//*[@id=\"village\"]/li[167]/a/div").click()
    time.sleep(1)



def click_digits(gata_no):
    for digit in str(gata_no):
        driver.find_element(By.XPATH,number_x_path_map[digit]).click()
        
        
def load_second_page(gata_no):
    click_digits(gata_no)
    driver.find_element(By.XPATH,"//*[@id=\"sbksn\"]/button[2]").click()
    time.sleep(1)
    multiple_gata=driver.find_elements(By.XPATH,"//input[@name=\"khata_number\"]")
    print(multiple_gata)
    print(len(multiple_gata))
    print(len(multiple_gata))
    if len(multiple_gata)>1:
        for i in range(0,len(multiple_gata)):
            multiple_gata[i].click()
            driver.find_element(By.XPATH,"//*[@id=\"sbksn\"]/button[1]").click()
            time.sleep(1)
            driver.find_element(By.XPATH,"//*[@id=\"capdiv\"]/button").click()
            gata_on_target_page=driver.find_elements(By.XPATH,"/html/body/center/form/table/tbody[2]/tr[3]/td[2]/font[1]/span/a")
            gata_on_target_page[0].click()
            
    else:
        multiple_gata[0].click()
        
         
            
            
        
        
        
    

    









load_first_page()
load_second_page(11)

    